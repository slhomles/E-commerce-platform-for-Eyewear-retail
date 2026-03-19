import openpyxl
wb = openpyxl.load_workbook('test_data/products_from_images.xlsx')
ws = wb.active
headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
print('Total columns:', len(row2))
print()
# Key columns to check
key_cols = {
    7: 'Sale Price',
    18: 'SKU',
    19: 'Color Name',
    20: 'Color Hex',
    21: 'Price Adjust',
    22: 'Initial Stock',
    23: 'Lens Width',
    26: 'Weight',
    27: 'THUMBNAIL URL',
    28: 'Gallery 1',
    29: 'Gallery 2',
}
print('=== FULL DATA ROW ===')
for i, (h, v) in enumerate(zip(headers, row2)):
    s = str(v)[:60] if v is not None else 'NONE'
    marker = ' <-- KEY' if i in key_cols else ''
    print(f'  [{i:2d}] Header={str(h)[:18]:18s} | Value={s}{marker}')
