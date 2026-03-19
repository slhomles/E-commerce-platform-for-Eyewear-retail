"""
Script đọc file product_import_template_FINAL.xlsx của người dùng,
điền thêm 14 sản phẩm vào (kèm URL ảnh Cloudinary) và lưu thành file mới.
"""

import os
import openpyxl
from openpyxl.styles import PatternFill, Alignment

# URLs of images already uploaded (reusing them to save time & API calls)
URLS = {
    "AF1037G-5S": ["https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855663/eyewear_import/AF1037G-5S_img0.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855664/eyewear_import/AF1037G-5S_img1.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855664/eyewear_import/AF1037G-5S_img2.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855665/eyewear_import/AF1037G-5S_img3.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855666/eyewear_import/AF1037G-5S_img4.jpg"],
    "AR2035T-1A": ["https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855666/eyewear_import/AR2035T-1A_img0.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855667/eyewear_import/AR2035T-1A_img1.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855668/eyewear_import/AR2035T-1A_img2.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855669/eyewear_import/AR2035T-1A_img3.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855669/eyewear_import/AR2035T-1A_img4.jpg"],
    "AU2087W-1S": ["https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855669/eyewear_import/AU2087W-1S_img0.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855670/eyewear_import/AU2087W-1S_img1.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855671/eyewear_import/AU2087W-1S_img2.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855671/eyewear_import/AU2087W-1S_img3.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855672/eyewear_import/AU2087W-1S_img4.jpg"],
    "AU2122N-5S": ["https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855672/eyewear_import/AU2122N-5S_img0.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855673/eyewear_import/AU2122N-5S_img1.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855673/eyewear_import/AU2122N-5S_img2.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855674/eyewear_import/AU2122N-5S_img3.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855675/eyewear_import/AU2122N-5S_img4.jpg"],
    "AU8005N-1A": ["https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855675/eyewear_import/AU8005N-1A_img0.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855676/eyewear_import/AU8005N-1A_img1.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855676/eyewear_import/AU8005N-1A_img2.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855677/eyewear_import/AU8005N-1A_img3.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855678/eyewear_import/AU8005N-1A_img4.jpg"],
    "DN1007B-5A": ["https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855678/eyewear_import/DN1007B-5A_img0.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855679/eyewear_import/DN1007B-5A_img1.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855679/eyewear_import/DN1007B-5A_img2.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855680/eyewear_import/DN1007B-5A_img3.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855681/eyewear_import/DN1007B-5A_img4.jpg"],
    "DN2008N-5A": ["https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855681/eyewear_import/DN2008N-5A_img0.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855682/eyewear_import/DN2008N-5A_img1.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855682/eyewear_import/DN2008N-5A_img2.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855683/eyewear_import/DN2008N-5A_img3.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855683/eyewear_import/DN2008N-5A_img4.jpg"],
    "HP2004B-5A": ["https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855684/eyewear_import/HP2004B-5A_img0.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855684/eyewear_import/HP2004B-5A_img1.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855685/eyewear_import/HP2004B-5A_img2.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855686/eyewear_import/HP2004B-5A_img3.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855686/eyewear_import/HP2004B-5A_img4.jpg"],
    "JU1025G-5S": ["https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855686/eyewear_import/JU1025G-5S_img0.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855687/eyewear_import/JU1025G-5S_img1.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855688/eyewear_import/JU1025G-5S_img2.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855688/eyewear_import/JU1025G-5S_img3.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855689/eyewear_import/JU1025G-5S_img4.jpg"],
    "LSA-123":    ["https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855689/eyewear_import/LSA-123_img0.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855690/eyewear_import/LSA-123_img1.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855691/eyewear_import/LSA-123_img2.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855691/eyewear_import/LSA-123_img3.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855692/eyewear_import/LSA-123_img4.jpg"],
    "MV2001N-5S": ["https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855693/eyewear_import/MV2001N-5S_img0.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855694/eyewear_import/MV2001N-5S_img1.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855695/eyewear_import/MV2001N-5S_img2.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855695/eyewear_import/MV2001N-5S_img3.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855696/eyewear_import/MV2001N-5S_img4.jpg"],
    "OB1015G-5A": ["https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855696/eyewear_import/OB1015G-5A_img0.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855697/eyewear_import/OB1015G-5A_img1.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855698/eyewear_import/OB1015G-5A_img2.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855698/eyewear_import/OB1015G-5A_img3.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855699/eyewear_import/OB1015G-5A_img4.jpg"],
    "OR1060X-5S": ["https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855699/eyewear_import/OR1060X-5S_img0.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855700/eyewear_import/OR1060X-5S_img1.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855701/eyewear_import/OR1060X-5S_img2.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855701/eyewear_import/OR1060X-5S_img3.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855702/eyewear_import/OR1060X-5S_img4.jpg"],
    "SR1010X-5A": ["https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855702/eyewear_import/SR1010X-5A_img0.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855703/eyewear_import/SR1010X-5A_img1.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855703/eyewear_import/SR1010X-5A_img2.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855704/eyewear_import/SR1010X-5A_img3.jpg", "https://res.cloudinary.com/dfk16iu6d/image/upload/v1773855705/eyewear_import/SR1010X-5A_img4.jpg"],
}

PRODUCT_META = {
    "AF1037G-5S": {"brand":"Afflelou","category":"Gong Kinh","type":"FRAME","base_price":1850000,"sale_price":1599000,"gender":"UNISEX","material":"Acetate","shape":"SQUARE","rim":"Full","hinge":"Standard","nose":"Fixed","frame_size":"54-17-145","style":"Classic","rx":"TRUE","prog":"FALSE","color_name":"Light Brown","color_hex":"#A0785A","stock":40,"lens_w":54,"bridge":17,"temple":145,"weight":24},
    "AR2035T-1A": {"brand":"Argon","category":"Gong Kinh","type":"FRAME","base_price":2200000,"sale_price":1890000,"gender":"MEN","material":"Titanium","shape":"RECTANGLE","rim":"Half","hinge":"Flexible","nose":"Adjustable","frame_size":"52-18-140","style":"Business","rx":"TRUE","prog":"TRUE","color_name":"Gunmetal","color_hex":"#797979","stock":35,"lens_w":52,"bridge":18,"temple":140,"weight":19},
    "AU2087W-1S": {"brand":"Aulola","category":"Kinh Mat","type":"FRAME","base_price":990000,"sale_price":799000,"gender":"WOMEN","material":"Acetate","shape":"OVAL","rim":"Full","hinge":"Standard","nose":"Fixed","frame_size":"53-19-140","style":"Fashion","rx":"FALSE","prog":"FALSE","color_name":"White","color_hex":"#F5F5F5","stock":50,"lens_w":53,"bridge":19,"temple":140,"weight":22},
    "AU2122N-5S": {"brand":"Aulola","category":"Gong Kinh","type":"FRAME","base_price":750000,"sale_price":"","gender":"KIDS","material":"TR90","shape":"RECTANGLE","rim":"Full","hinge":"Spring","nose":"Silicone","frame_size":"46-15-125","style":"Kids","rx":"TRUE","prog":"FALSE","color_name":"Navy","color_hex":"#1B2A4A","stock":60,"lens_w":46,"bridge":15,"temple":125,"weight":16},
    "AU8005N-1A": {"brand":"Aulola Premium","category":"Gong Kinh","type":"FRAME","base_price":3200000,"sale_price":2700000,"gender":"MEN","material":"Titanium","shape":"RECTANGLE","rim":"Full","hinge":"Flexible","nose":"Adjustable","frame_size":"55-17-148","style":"Luxury","rx":"TRUE","prog":"TRUE","color_name":"Black","color_hex":"#1A1A1A","stock":20,"lens_w":55,"bridge":17,"temple":148,"weight":18},
    "DN1007B-5A": {"brand":"Dior Inspired","category":"Kinh Mat","type":"FRAME","base_price":1350000,"sale_price":1100000,"gender":"WOMEN","material":"Acetate","shape":"ROUND","rim":"Full","hinge":"Standard","nose":"Fixed","frame_size":"52-20-145","style":"Vintage","rx":"FALSE","prog":"FALSE","color_name":"Tortoise Brown","color_hex":"#7B4F2E","stock":45,"lens_w":52,"bridge":20,"temple":145,"weight":25},
    "DN2008N-5A": {"brand":"Dior Inspired","category":"Kinh Mat","type":"FRAME","base_price":1550000,"sale_price":1299000,"gender":"WOMEN","material":"Acetate","shape":"CAT_EYE","rim":"Full","hinge":"Standard","nose":"Fixed","frame_size":"55-16-140","style":"Fashion","rx":"FALSE","prog":"FALSE","color_name":"Black","color_hex":"#1A1A1A","stock":38,"lens_w":55,"bridge":16,"temple":140,"weight":23},
    "HP2004B-5A": {"brand":"Hoya Prestige","category":"Gong Kinh","type":"FRAME","base_price":1750000,"sale_price":"","gender":"MEN","material":"Stainless Steel","shape":"SQUARE","rim":"Full","hinge":"Standard","nose":"Adjustable","frame_size":"54-16-142","style":"Classic","rx":"TRUE","prog":"TRUE","color_name":"Blue Steel","color_hex":"#4A7FA5","stock":30,"lens_w":54,"bridge":16,"temple":142,"weight":21},
    "JU1025G-5S": {"brand":"Juliette","category":"Kinh Mat","type":"FRAME","base_price":2800000,"sale_price":2390000,"gender":"WOMEN","material":"Acetate","shape":"OVAL","rim":"Full","hinge":"Gold Hinge","nose":"Fixed","frame_size":"56-18-145","style":"Luxury","rx":"FALSE","prog":"FALSE","color_name":"Gold Havana","color_hex":"#C5942A","stock":25,"lens_w":56,"bridge":18,"temple":145,"weight":28},
    "LSA-123":    {"brand":"Lensara","category":"Gong Kinh","type":"FRAME","base_price":1200000,"sale_price":980000,"gender":"UNISEX","material":"Titanium","shape":"RECTANGLE","rim":"Rimless","hinge":"Flexible","nose":"Adjustable","frame_size":"53-17-140","style":"Minimalist","rx":"TRUE","prog":"TRUE","color_name":"Silver","color_hex":"#C0C0C0","stock":35,"lens_w":53,"bridge":17,"temple":140,"weight":14},
    "MV2001N-5S": {"brand":"Maverick Sport","category":"Kinh Mat","type":"FRAME","base_price":1650000,"sale_price":1420000,"gender":"MEN","material":"TR90","shape":"WRAP_AROUND","rim":"Full","hinge":"Rubber","nose":"Silicone","frame_size":"58-18-130","style":"Sport","rx":"FALSE","prog":"FALSE","color_name":"Matte Black","color_hex":"#2C2C2C","stock":55,"lens_w":58,"bridge":18,"temple":130,"weight":26},
    "OB1015G-5A": {"brand":"Obsidian","category":"Kinh Mat","type":"FRAME","base_price":1450000,"sale_price":1199000,"gender":"UNISEX","material":"Acetate","shape":"BROWLINE","rim":"Half","hinge":"Standard","nose":"Fixed","frame_size":"51-21-145","style":"Retro","rx":"FALSE","prog":"FALSE","color_name":"Gold Brown","color_hex":"#A67B5B","stock":42,"lens_w":51,"bridge":21,"temple":145,"weight":27},
    "OR1060X-5S": {"brand":"Orion","category":"Kinh Mat","type":"FRAME","base_price":1300000,"sale_price":1099000,"gender":"MEN","material":"Stainless Steel","shape":"AVIATOR","rim":"Half","hinge":"Standard","nose":"Adjustable","frame_size":"58-14-135","style":"Classic","rx":"FALSE","prog":"FALSE","color_name":"Gold","color_hex":"#D4AF37","stock":48,"lens_w":58,"bridge":14,"temple":135,"weight":22},
    "SR1010X-5A": {"brand":"Sorrento","category":"Gong Kinh","type":"FRAME","base_price":2100000,"sale_price":1790000,"gender":"UNISEX","material":"Acetate","shape":"SQUARE","rim":"Full","hinge":"Spring","nose":"Adjustable","frame_size":"53-18-145","style":"Modern","rx":"TRUE","prog":"FALSE","color_name":"Tortoise","color_hex":"#6B3A2A","stock":33,"lens_w":53,"bridge":18,"temple":145,"weight":23},
}

INPUT_FILE = os.path.join(os.path.dirname(__file__), "product_import_template_FINAL.xlsx")
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "product_import_template_FINAL_FILLED.xlsx")

def write_cell(ws, row_idx, col_num, value, fill=None):
    cell = ws.cell(row=row_idx, column=col_num, value=value)
    if fill:
        cell.fill = fill
    cell.alignment = Alignment(vertical="center")

def main():
    print(f"Loading {INPUT_FILE}...")
    wb = openpyxl.load_workbook(INPUT_FILE)
    if "Products" in wb.sheetnames:
        ws = wb["Products"]
    else:
        ws = wb.active
        
    print(f"Active sheet: {ws.title}")

    # Xoa cac dong du lieu cu (tu row 2 tro xuong)
    ws.delete_rows(2, ws.max_row)

    row_idx = 2
    for sku, meta in PRODUCT_META.items():
        name = meta.get("name", "Gong Kinh " + sku)
        fill = PatternFill("solid", fgColor="EBF3FB" if row_idx % 2 == 0 else "FFFFFF")
        
        write_cell(ws, row_idx, 1, name, fill)
        write_cell(ws, row_idx, 2, name.lower().replace(" ", "-"), fill)
        write_cell(ws, row_idx, 3, f"San pham {name} chat luong cao", fill)
        write_cell(ws, row_idx, 4, meta["brand"], fill)
        write_cell(ws, row_idx, 5, meta["category"], fill)
        write_cell(ws, row_idx, 6, meta["type"], fill)
        write_cell(ws, row_idx, 7, meta["base_price"], fill)
        write_cell(ws, row_idx, 8, meta["sale_price"], fill)
        write_cell(ws, row_idx, 9, meta["gender"], fill)
        write_cell(ws, row_idx, 10, meta["material"], fill)
        write_cell(ws, row_idx, 11, meta["shape"], fill)
        write_cell(ws, row_idx, 12, meta["rim"], fill)
        write_cell(ws, row_idx, 13, meta["hinge"], fill)
        write_cell(ws, row_idx, 14, meta["nose"], fill)
        write_cell(ws, row_idx, 15, meta["frame_size"], fill)
        write_cell(ws, row_idx, 16, meta["style"], fill)
        write_cell(ws, row_idx, 17, meta["rx"], fill)
        write_cell(ws, row_idx, 18, meta["prog"], fill)
        write_cell(ws, row_idx, 19, sku, fill)
        write_cell(ws, row_idx, 20, meta["color_name"], fill)
        write_cell(ws, row_idx, 21, meta["color_hex"], fill)
        write_cell(ws, row_idx, 22, 0, fill) # Price Adjust
        write_cell(ws, row_idx, 23, meta["stock"], fill)
        write_cell(ws, row_idx, 24, meta["lens_w"], fill)
        write_cell(ws, row_idx, 25, meta["bridge"], fill)
        write_cell(ws, row_idx, 26, meta["temple"], fill)
        write_cell(ws, row_idx, 27, meta["weight"], fill)
        
        urls = URLS.get(sku, ["","","","",""])
        write_cell(ws, row_idx, 28, urls[0], fill)
        write_cell(ws, row_idx, 29, urls[1], fill)
        write_cell(ws, row_idx, 30, urls[2], fill)
        write_cell(ws, row_idx, 31, urls[3], fill)
        write_cell(ws, row_idx, 32, urls[4], fill)
        
        row_idx += 1

    wb.save(OUTPUT_FILE)
    print(f"Saved to {OUTPUT_FILE} with {row_idx-2} lines full of data!")

if __name__ == "__main__":
    main()
