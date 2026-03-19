"""
Script tao Excel dung cot bang cach ghi tung o theo ten cot (khong dung array index).
"""
import os
import cloudinary
import cloudinary.uploader
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

cloudinary.config(
    cloud_name="dfk16iu6d",
    api_key="731262262332786",
    api_secret="SEQfYuJpR3rSD4pmElGQoKleAQE",
    secure=True
)

IMAGES_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "images"))
OUTPUT_EXCEL = os.path.join(os.path.dirname(__file__), "import_ready.xlsx")
SUPPORTED_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}

# Map field -> column number (1-based, A=1 ... AF=32)
COL = {
    "name": 1, "slug": 2, "description": 3, "brand": 4, "category": 5,
    "type": 6, "base_price": 7, "sale_price": 8, "gender": 9,
    "material": 10, "shape": 11, "rim": 12, "hinge": 13, "nose": 14,
    "frame_size": 15, "style": 16, "rx": 17, "prog": 18, "sku": 19,
    "color_name": 20, "color_hex": 21, "price_adj": 22, "stock": 23,
    "lens_w": 24, "bridge": 25, "temple": 26, "weight": 27,
    "thumb_url": 28, "gal_1": 29, "gal_2": 30, "gal_3": 31, "gal_4": 32,
}

PRODUCT_META = {
    "AF1037G-5S": {"name":"Kinh Gong Acetate AF1037G","brand":"Afflelou","category":"Gong Kinh","type":"FRAME","base_price":1850000,"sale_price":1599000,"gender":"UNISEX","material":"Acetate","shape":"SQUARE","rim":"Full","hinge":"Standard","nose":"Fixed","frame_size":"54-17-145","style":"Classic","rx":"TRUE","prog":"FALSE","color_name":"Light Brown","color_hex":"#A0785A","stock":40,"lens_w":54,"bridge":17,"temple":145,"weight":24},
    "AR2035T-1A": {"name":"Gong Kinh Titan Mong AR2035T","brand":"Argon","category":"Gong Kinh","type":"FRAME","base_price":2200000,"sale_price":1890000,"gender":"MEN","material":"Titanium","shape":"RECTANGLE","rim":"Half","hinge":"Flexible","nose":"Adjustable","frame_size":"52-18-140","style":"Business","rx":"TRUE","prog":"TRUE","color_name":"Gunmetal","color_hex":"#797979","stock":35,"lens_w":52,"bridge":18,"temple":140,"weight":19},
    "AU2087W-1S": {"name":"Kinh Mat Oval AU2087W","brand":"Aulola","category":"Kinh Mat","type":"FRAME","base_price":990000,"sale_price":799000,"gender":"WOMEN","material":"Acetate","shape":"OVAL","rim":"Full","hinge":"Standard","nose":"Fixed","frame_size":"53-19-140","style":"Fashion","rx":"FALSE","prog":"FALSE","color_name":"White","color_hex":"#F5F5F5","stock":50,"lens_w":53,"bridge":19,"temple":140,"weight":22},
    "AU2122N-5S": {"name":"Kinh Gong Nhua AU2122N","brand":"Aulola","category":"Gong Kinh","type":"FRAME","base_price":750000,"sale_price":"","gender":"KIDS","material":"TR90","shape":"RECTANGLE","rim":"Full","hinge":"Spring","nose":"Silicone","frame_size":"46-15-125","style":"Kids","rx":"TRUE","prog":"FALSE","color_name":"Navy","color_hex":"#1B2A4A","stock":60,"lens_w":46,"bridge":15,"temple":125,"weight":16},
    "AU8005N-1A": {"name":"Gong Kinh Cao Cap AU8005N","brand":"Aulola Premium","category":"Gong Kinh","type":"FRAME","base_price":3200000,"sale_price":2700000,"gender":"MEN","material":"Titanium","shape":"RECTANGLE","rim":"Full","hinge":"Flexible","nose":"Adjustable","frame_size":"55-17-148","style":"Luxury","rx":"TRUE","prog":"TRUE","color_name":"Black","color_hex":"#1A1A1A","stock":20,"lens_w":55,"bridge":17,"temple":148,"weight":18},
    "DN1007B-5A": {"name":"Kinh Mat Tron DN1007B","brand":"Dior Inspired","category":"Kinh Mat","type":"FRAME","base_price":1350000,"sale_price":1100000,"gender":"WOMEN","material":"Acetate","shape":"ROUND","rim":"Full","hinge":"Standard","nose":"Fixed","frame_size":"52-20-145","style":"Vintage","rx":"FALSE","prog":"FALSE","color_name":"Tortoise Brown","color_hex":"#7B4F2E","stock":45,"lens_w":52,"bridge":20,"temple":145,"weight":25},
    "DN2008N-5A": {"name":"Kinh Mat Cat Eye DN2008N","brand":"Dior Inspired","category":"Kinh Mat","type":"FRAME","base_price":1550000,"sale_price":1299000,"gender":"WOMEN","material":"Acetate","shape":"CAT_EYE","rim":"Full","hinge":"Standard","nose":"Fixed","frame_size":"55-16-140","style":"Fashion","rx":"FALSE","prog":"FALSE","color_name":"Black","color_hex":"#1A1A1A","stock":38,"lens_w":55,"bridge":16,"temple":140,"weight":23},
    "HP2004B-5A": {"name":"Gong Kinh Vuong HP2004B","brand":"Hoya Prestige","category":"Gong Kinh","type":"FRAME","base_price":1750000,"sale_price":"","gender":"MEN","material":"Stainless Steel","shape":"SQUARE","rim":"Full","hinge":"Standard","nose":"Adjustable","frame_size":"54-16-142","style":"Classic","rx":"TRUE","prog":"TRUE","color_name":"Blue Steel","color_hex":"#4A7FA5","stock":30,"lens_w":54,"bridge":16,"temple":142,"weight":21},
    "JU1025G-5S": {"name":"Kinh Mat Luxe JU1025G","brand":"Juliette","category":"Kinh Mat","type":"FRAME","base_price":2800000,"sale_price":2390000,"gender":"WOMEN","material":"Acetate","shape":"OVAL","rim":"Full","hinge":"Gold Hinge","nose":"Fixed","frame_size":"56-18-145","style":"Luxury","rx":"FALSE","prog":"FALSE","color_name":"Gold Havana","color_hex":"#C5942A","stock":25,"lens_w":56,"bridge":18,"temple":145,"weight":28},
    "LSA-123":    {"name":"Gong Kinh Rimless LSA-123","brand":"Lensara","category":"Gong Kinh","type":"FRAME","base_price":1200000,"sale_price":980000,"gender":"UNISEX","material":"Titanium","shape":"RECTANGLE","rim":"Rimless","hinge":"Flexible","nose":"Adjustable","frame_size":"53-17-140","style":"Minimalist","rx":"TRUE","prog":"TRUE","color_name":"Silver","color_hex":"#C0C0C0","stock":35,"lens_w":53,"bridge":17,"temple":140,"weight":14},
    "MV2001N-5S": {"name":"Kinh Mat Sport MV2001N","brand":"Maverick Sport","category":"Kinh Mat","type":"FRAME","base_price":1650000,"sale_price":1420000,"gender":"MEN","material":"TR90","shape":"WRAP_AROUND","rim":"Full","hinge":"Rubber","nose":"Silicone","frame_size":"58-18-130","style":"Sport","rx":"FALSE","prog":"FALSE","color_name":"Matte Black","color_hex":"#2C2C2C","stock":55,"lens_w":58,"bridge":18,"temple":130,"weight":26},
    "OB1015G-5A": {"name":"Kinh Mat Clubmaster OB1015G","brand":"Obsidian","category":"Kinh Mat","type":"FRAME","base_price":1450000,"sale_price":1199000,"gender":"UNISEX","material":"Acetate","shape":"BROWLINE","rim":"Half","hinge":"Standard","nose":"Fixed","frame_size":"51-21-145","style":"Retro","rx":"FALSE","prog":"FALSE","color_name":"Gold Brown","color_hex":"#A67B5B","stock":42,"lens_w":51,"bridge":21,"temple":145,"weight":27},
    "OR1060X-5S": {"name":"Kinh Mat Phi Cong OR1060X","brand":"Orion","category":"Kinh Mat","type":"FRAME","base_price":1300000,"sale_price":1099000,"gender":"MEN","material":"Stainless Steel","shape":"AVIATOR","rim":"Half","hinge":"Standard","nose":"Adjustable","frame_size":"58-14-135","style":"Classic","rx":"FALSE","prog":"FALSE","color_name":"Gold","color_hex":"#D4AF37","stock":48,"lens_w":58,"bridge":14,"temple":135,"weight":22},
    "SR1010X-5A": {"name":"Gong Kinh Vuong Premium SR1010X","brand":"Sorrento","category":"Gong Kinh","type":"FRAME","base_price":2100000,"sale_price":1790000,"gender":"UNISEX","material":"Acetate","shape":"SQUARE","rim":"Full","hinge":"Spring","nose":"Adjustable","frame_size":"53-18-145","style":"Modern","rx":"TRUE","prog":"FALSE","color_name":"Tortoise","color_hex":"#6B3A2A","stock":33,"lens_w":53,"bridge":18,"temple":145,"weight":23},
}

def get_sorted_images(folder_path, max_count=5):
    files = []
    for f in sorted(os.listdir(folder_path)):
        if os.path.splitext(f)[1].lower() in SUPPORTED_EXTS:
            files.append(os.path.join(folder_path, f))
        if len(files) >= max_count:
            break
    return files

def upload_image(local_path, public_id):
    try:
        r = cloudinary.uploader.upload(local_path, public_id=public_id,
                                       folder="eyewear_import", overwrite=True,
                                       resource_type="image")
        url = r.get("secure_url", "")
        print(f"  OK {os.path.basename(local_path)}")
        return url
    except Exception as e:
        print(f"  ERR {e}")
        return ""

def wcell(ws, row, field, value, fill):
    """Write value to the correct named column."""
    c = ws.cell(row=row, column=COL[field], value=value)
    c.fill = fill
    c.alignment = Alignment(vertical="center")

def main():
    subfolders = sorted([f for f in os.listdir(IMAGES_ROOT)
                         if os.path.isdir(os.path.join(IMAGES_ROOT, f))])
    print(f"Tim thay {len(subfolders)} san pham")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    hf = Font(color="FFFFFF", bold=True, size=9)
    h_fill   = PatternFill("solid", fgColor="1F4E79")
    img_fill = PatternFill("solid", fgColor="2E75B6")
    req_fill = PatternFill("solid", fgColor="C55A11")
    req_keys = {"name", "brand", "category", "base_price", "sku", "thumb_url"}
    img_keys = {"thumb_url", "gal_1", "gal_2", "gal_3", "gal_4"}

    HEADERS = {
        "name": "Name (*)", "slug": "Slug", "description": "Mo Ta",
        "brand": "Brand (*)", "category": "Category (*)", "type": "Type",
        "base_price": "Gia Goc (*)", "sale_price": "Gia Sale",
        "gender": "Gioi Tinh", "material": "Chat Lieu", "shape": "Kieu Dang",
        "rim": "Vien Gong", "hinge": "Ban Le", "nose": "De Mui",
        "frame_size": "Kich Thuoc", "style": "Phong Cach",
        "rx": "Ho Tro Ke Don?", "prog": "Luy Tien?", "sku": "SKU (*)",
        "color_name": "Mau Sac", "color_hex": "Ma Hex",
        "price_adj": "Dieu Chinh Gia", "stock": "Ton Kho Ban Dau",
        "lens_w": "Rong Kinh (mm)", "bridge": "Cau Mui (mm)",
        "temple": "Gong (mm)", "weight": "Can Nang (g)",
        "thumb_url": "THUMBNAIL URL [AB] (*)", "gal_1": "Gallery 1 URL [AC]",
        "gal_2": "Gallery 2 URL [AD]", "gal_3": "Gallery 3 URL [AE]",
        "gal_4": "Gallery 4 URL [AF]",
    }

    for field, col_num in COL.items():
        c = ws.cell(row=1, column=col_num, value=HEADERS.get(field, field))
        c.fill = img_fill if field in img_keys else (req_fill if field in req_keys else h_fill)
        c.font = hf
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_idx = 2
    for fi, folder_name in enumerate(subfolders, start=1):
        sku  = folder_name
        meta = PRODUCT_META.get(sku, {})
        name = meta.get("name", sku)
        print(f"\n[{fi}/{len(subfolders)}] {sku}")

        pics = get_sorted_images(os.path.join(IMAGES_ROOT, sku), max_count=5)
        urls = [upload_image(p, f"{sku}_img{i}") for i, p in enumerate(pics)]
        while len(urls) < 5:
            urls.append("")

        fill = PatternFill("solid", fgColor="EBF3FB" if row_idx % 2 == 0 else "FFFFFF")

        wcell(ws, row_idx, "name",        name, fill)
        wcell(ws, row_idx, "slug",        name.lower().replace(" ", "-"), fill)
        wcell(ws, row_idx, "description", f"San pham {name} chat luong cao, thiet ke hien dai", fill)
        wcell(ws, row_idx, "brand",       meta.get("brand", "Eyewear Brand"), fill)
        wcell(ws, row_idx, "category",    meta.get("category", "Kinh Mat"), fill)
        wcell(ws, row_idx, "type",        meta.get("type", "FRAME"), fill)
        wcell(ws, row_idx, "base_price",  meta.get("base_price", 1000000), fill)
        sp = meta.get("sale_price", "")
        wcell(ws, row_idx, "sale_price",  sp if sp != "" else None, fill)
        wcell(ws, row_idx, "gender",      meta.get("gender", "UNISEX"), fill)
        wcell(ws, row_idx, "material",    meta.get("material", "Acetate"), fill)
        wcell(ws, row_idx, "shape",       meta.get("shape", "RECTANGLE"), fill)
        wcell(ws, row_idx, "rim",         meta.get("rim", "Full"), fill)
        wcell(ws, row_idx, "hinge",       meta.get("hinge", "Standard"), fill)
        wcell(ws, row_idx, "nose",        meta.get("nose", "Adjustable"), fill)
        wcell(ws, row_idx, "frame_size",  meta.get("frame_size", "52-18-140"), fill)
        wcell(ws, row_idx, "style",       meta.get("style", "Classic"), fill)
        wcell(ws, row_idx, "rx",          meta.get("rx", "FALSE"), fill)
        wcell(ws, row_idx, "prog",        meta.get("prog", "FALSE"), fill)
        wcell(ws, row_idx, "sku",         sku, fill)
        wcell(ws, row_idx, "color_name",  meta.get("color_name", "Black"), fill)
        wcell(ws, row_idx, "color_hex",   meta.get("color_hex", "#1A1A1A"), fill)
        wcell(ws, row_idx, "price_adj",   0, fill)
        wcell(ws, row_idx, "stock",       meta.get("stock", 50), fill)
        wcell(ws, row_idx, "lens_w",      meta.get("lens_w", 52), fill)
        wcell(ws, row_idx, "bridge",      meta.get("bridge", 18), fill)
        wcell(ws, row_idx, "temple",      meta.get("temple", 140), fill)
        wcell(ws, row_idx, "weight",      meta.get("weight", 22), fill)
        wcell(ws, row_idx, "thumb_url",   urls[0], fill)
        wcell(ws, row_idx, "gal_1",       urls[1], fill)
        wcell(ws, row_idx, "gal_2",       urls[2], fill)
        wcell(ws, row_idx, "gal_3",       urls[3], fill)
        wcell(ws, row_idx, "gal_4",       urls[4], fill)

        row_idx += 1

    # Column widths
    wmap = {"name":25,"description":40,"brand":14,"category":12,"sku":16,
            "color_hex":10,"thumb_url":55,"gal_1":55,"gal_2":55,"gal_3":55,"gal_4":55}
    for field, col_num in COL.items():
        ws.column_dimensions[get_column_letter(col_num)].width = wmap.get(field, 14)
    ws.row_dimensions[1].height = 32
    for r in range(2, row_idx):
        ws.row_dimensions[r].height = 22
    ws.freeze_panes = "A2"

    wb.save(OUTPUT_EXCEL)
    print(f"\n=== XONG! {row_idx-2} san pham -> {OUTPUT_EXCEL} ===\n")
    print("Xac nhan mapping cot:")
    for field, col_num in COL.items():
        print(f"  {field:12s} -> Excel col {col_num:2d} ({get_column_letter(col_num)}) -> POI index {col_num-1:2d}")

if __name__ == "__main__":
    main()
