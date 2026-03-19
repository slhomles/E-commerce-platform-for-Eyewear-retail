"""
Script tao file Excel mau - Import San Pham Thong Minh

CAU TRUC COT:
  0  (A)  Name
  1  (B)  Slug
  2  (C)  Description
  3  (D)  Brand
  4  (E)  Category
  5  (F)  Type  (FRAME/LENS/SERVICE)
  6  (G)  Base Price
  7  (H)  Sale Price
  8  (I)  Gender (MEN/WOMEN/UNISEX/KIDS)
  9  (J)  Frame Material
  10 (K)  Frame Shape
  11 (L)  Rim Type
  12 (M)  Hinge Type
  13 (N)  Nose Pad Type
  14 (O)  Frame Size   (52-18-140)
  15 (P)  Style
  16 (Q)  Ke Don?      (TRUE/FALSE)
  17 (R)  Luy Tien?    (TRUE/FALSE)
  18 (S)  SKU          *** BAT BUOC ***
  19 (T)  Color Name
  20 (U)  Color Hex    (#rrggbb)
  21 (V)  Price Adjustment
  22 (W)  Initial Stock
  23 (X)  Lens Width (mm)
  24 (Y)  Bridge Width (mm)
  25 (Z)  Temple Length (mm)
  26 (AA) Weight (g)
  27 (AB) THUMBNAIL    *** ANH CHINH - Place in Cell ***
  28 (AC) Gallery 1    *** ANH PHU 1 - Place in Cell ***
  29 (AD) Gallery 2    *** ANH PHU 2 - Place in Cell ***
  30 (AE) Gallery 3    *** ANH PHU 3 - Place in Cell ***
  31 (AF) Gallery 4    *** ANH PHU 4 - Place in Cell ***
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os


def create_product_template(filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    # ====== HEADERS (32 cot: A -> AF) ======
    headers = [
        "Name",               # A  - 0
        "Slug",               # B  - 1
        "Description",        # C  - 2
        "Brand",              # D  - 3
        "Category",           # E  - 4
        "Type",               # F  - 5
        "Base Price",         # G  - 6
        "Sale Price",         # H  - 7
        "Gender",             # I  - 8
        "Frame Material",     # J  - 9
        "Frame Shape",        # K  - 10
        "Rim Type",           # L  - 11
        "Hinge Type",         # M  - 12
        "Nose Pad Type",      # N  - 13
        "Frame Size",         # O  - 14
        "Style",              # P  - 15
        "Ke Don?",            # Q  - 16
        "Luy Tien?",          # R  - 17
        "SKU (*)",            # S  - 18 BAT BUOC
        "Color Name",         # T  - 19
        "Color Hex",          # U  - 20
        "Price Adjust",       # V  - 21
        "Initial Stock",      # W  - 22
        "Lens Width (mm)",    # X  - 23
        "Bridge Width (mm)",  # Y  - 24
        "Temple Len (mm)",    # Z  - 25
        "Weight (g)",         # AA - 26
        "THUMBNAIL URL [AB] (*)", # AB - 27 *** URL ảnh chính ***
        "Gallery 1 URL [AC]",      # AC - 28
        "Gallery 2 URL [AD]",      # AD - 29
        "Gallery 3 URL [AE]",      # AE - 30
        "Gallery 4 URL [AF]",      # AF - 31
    ]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    img_fill    = PatternFill("solid", fgColor="2E75B6")
    req_fill    = PatternFill("solid", fgColor="C55A11")
    header_font = Font(color="FFFFFF", bold=True, size=9)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = img_fill if col_idx >= 28 else header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # To cam cac cot bat buoc: Name, Base Price, Brand, Category, SKU
    for col_idx in [1, 7, 4, 5, 19]:
        ws.cell(row=1, column=col_idx).fill = req_fill

    # ====== DU LIEU MAU ======
    sample_rows = [
        [
            "Kinh Mat Gucci Cao Cap",
            "",
            "Kinh mat sang trong, chong UV400",
            "Gucci",
            "Kinh Mat",
            "FRAME",
            2500000, 2100000,
            "UNISEX",
            "Acetate", "OVAL", "Full", "Standard", "Adjustable",
            "54-18-145", "Luxury",
            "FALSE", "FALSE",
            "GUC-001-BLK",
            "Den Bong", "#1A1A1A",
            0, 25,
            54, 18, 145, 28,
            # AB-AF: de trong, ban tu chen anh bang Place in Cell
        ],
        [
            "Gong Kinh RayBan Classic",
            "",
            "Gong kinh nhe, thiet ke co dien",
            "Ray-Ban",
            "Gong Kinh",
            "FRAME",
            1200000, 980000,
            "MEN",
            "Titanium", "RECTANGLE", "Half", "Flexible", "Silicone",
            "52-17-140", "Classic",
            "TRUE", "TRUE",
            "RB-001-GLD",
            "Vang Mo", "#C5A028",
            0, 30,
            52, 17, 140, 22,
        ],
    ]

    even_fill = PatternFill("solid", fgColor="EBF3FB")
    odd_fill  = PatternFill("solid", fgColor="FFFFFF")

    for row_idx, row_data in enumerate(sample_rows, start=2):
        fill = even_fill if row_idx % 2 == 0 else odd_fill
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")

    # ====== DO RONG COT ======
    col_widths = [
        30, 18, 40, 14, 14, 10, 14, 14, 10, 14,  # A-J
        14, 10, 12, 14, 15, 12, 10, 12, 16, 14,  # K-T
        12, 14, 14, 14, 15, 16, 12,               # U-AA
        24, 20, 20, 20, 20,                        # AB-AF (cot anh)
    ]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 32
    for r in range(2, 2 + len(sample_rows)):
        ws.row_dimensions[r].height = 80

    ws.freeze_panes = "A2"

    # ====== SHEET HUONG DAN ======
    ws_guide = wb.create_sheet("HUONG DAN")
    ws_guide["A1"] = "HUONG DAN CHEN ANH (5 anh cho moi san pham)"
    ws_guide["A1"].font = Font(bold=True, size=14, color="1F4E79")

    steps = [
        ("Cach dung:", "Dan URL anh cong khai vao cot AB (Thumbnail) va AC-AF (Gallery 1-4)"),
        ("URL hop le:", "Bat dau bang http:// hoac https:// (vi du tu Cloudinary, Imgur, Google Drive CDN...)"),
        ("Buoc 1:", "Upload anh len bat ky dich vu luu tru anh cong khai (Cloudinary, Imgur, ...)"),
        ("Buoc 2:", "Sao chep URL truc tiep cua anh (phai la URL truc tiep den file .jpg/.png)"),
        ("Buoc 3:", "Dan URL vao o AB2 cho Thumbnail san pham 1"),
        ("Buoc 4:", "Dan URL vao AC2, AD2, AE2, AF2 cho cac anh gallery cua san pham 1"),
        ("Buoc 5:", "Lam tuong tu cho hang 3, 4... (moi hang = 1 san pham)"),
        ("Buoc 6:", "Luu file -> Upload qua Admin -> Import Excel"),
        ("LUU Y 1:", "Neu khong co anh gallery, de trong cac o AC-AF cung khong sao"),
        ("LUU Y 2:", "URL Cloudinary vi du: https://res.cloudinary.com/abc/image/upload/v123/photo.jpg"),
        ("LUU Y 3:", "SKU la bat buoc va phai khac nhau. SKU da ton tai => cap nhat san pham."),
    ]
    for r, (step, desc) in enumerate(steps, start=3):
        ws_guide.cell(row=r, column=1, value=step).font = Font(bold=True, color="C55A11")
        ws_guide.cell(row=r, column=2, value=desc)
    ws_guide.column_dimensions["A"].width = 14
    ws_guide.column_dimensions["B"].width = 90

    wb.save(filename)
    print(f"File da tao: {filename}")
    print(f"   - {len(headers)} cot (A -> AF)")
    print(f"   - 5 o anh: AB (Thumbnail) + AC,AD,AE,AF (Gallery 1-4)")
    print(f"   - {len(sample_rows)} hang du lieu mau")


if __name__ == "__main__":
    os.makedirs("test_data", exist_ok=True)
    output_path = os.path.join("test_data", "product_import_template_FINAL.xlsx")
    create_product_template(output_path)
